import openpyxl
from django.contrib.auth import get_user_model
from django.http import HttpResponse
from django.utils import timezone
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from rest_framework import permissions, status, viewsets
from rest_framework.decorators import action
from rest_framework.response import Response

from apps.accounts.permissions import CanViewDeliveries
from apps.notifications.models import Notification

from .models import Delivery, DeliveryLocationLog
from .serializers import (
    CourierLocationSerializer,
    DeliveryCreateSerializer,
    DeliveryDetailSerializer,
    DeliveryLocationLogSerializer,
    DeliverySerializer,
)

User = get_user_model()


class DeliveryViewSet(viewsets.ModelViewSet):
    queryset = Delivery.objects.none()
    permission_classes = [CanViewDeliveries]

    def get_serializer_class(self):
        if self.action == 'create':
            return DeliveryCreateSerializer
        if self.action in ['retrieve']:
            return DeliveryDetailSerializer
        return DeliverySerializer

    def get_queryset(self):
        if getattr(self, 'swagger_fake_view', False):
            return Delivery.objects.none()
        user = self.request.user
        if user.is_anonymous:
            return Delivery.objects.none()
        if user.is_super_admin:
            return Delivery.objects.all()
        qs = Delivery.objects.select_related('order__pharmacy', 'courier')
        if user.role == 'pharmacy':
            if hasattr(user, 'pharmacy_profile'):
                qs = qs.filter(order__pharmacy=user.pharmacy_profile)
            else:
                qs = qs.none()
        return qs

    @action(detail=True, methods=['post'])
    def update_location(self, request, pk=None):
        serializer = CourierLocationSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        delivery = self.get_object()
        delivery.courier_lat = serializer.validated_data['latitude']
        delivery.courier_lng = serializer.validated_data['longitude']
        delivery.courier_location_updated_at = timezone.now()
        delivery.save(update_fields=['courier_lat', 'courier_lng', 'courier_location_updated_at'])

        DeliveryLocationLog.objects.create(
            delivery=delivery,
            courier=request.user,
            latitude=serializer.validated_data['latitude'],
            longitude=serializer.validated_data['longitude'],
            accuracy=serializer.validated_data.get('accuracy'),
            speed=serializer.validated_data.get('speed'),
            bearing=serializer.validated_data.get('bearing'),
            battery_level=serializer.validated_data.get('battery_level'),
        )

        return Response({'message': 'Joylashuv yangilandi'})

    @action(detail=True, methods=['post'])
    def assign_courier(self, request, pk=None):
        courier_id = request.data.get('courier_id')
        if not courier_id:
            return Response({'error': 'Kuryer ID si talab qilinadi'}, status=status.HTTP_400_BAD_REQUEST)
        delivery = self.get_object()
        try:
            courier = User.objects.get(id=courier_id, role='operator')
        except User.DoesNotExist:
            return Response({'error': 'Kuryer topilmadi'}, status=status.HTTP_404_NOT_FOUND)
        delivery.courier = courier
        delivery.status = 'assigned'
        delivery.assigned_at = timezone.now()
        delivery.save()

        pharmacy_user = delivery.order.pharmacy.user if delivery.order.pharmacy and delivery.order.pharmacy.user else None
        if pharmacy_user and pharmacy_user.is_active:
            Notification.objects.create(
                user=pharmacy_user,
                type='system',
                title='Yetkazib berishga kuryer biriktirildi',
                message=f'{delivery.order.order_number} - kuryer: {courier.get_full_name() or courier.login}',
                link=f'/pharmacy/orders/{delivery.order.id}',
            )

        return Response(DeliverySerializer(delivery, context={'request': request}).data)

    @action(detail=True, methods=['post'])
    def update_status(self, request, pk=None):
        new_status = request.data.get('status')
        if new_status not in dict(Delivery.STATUS_CHOICES):
            return Response({'error': 'Noto\'g\'ri holat'}, status=status.HTTP_400_BAD_REQUEST)
        delivery = self.get_object()
        delivery.status = new_status
        if new_status == 'picked':
            delivery.picked_at = timezone.now()
        elif new_status == 'delivered':
            delivery.delivered_at = timezone.now()
        delivery.save()

        status_labels = dict(Delivery.STATUS_CHOICES)
        pharmacy_user = delivery.order.pharmacy.user if delivery.order.pharmacy and delivery.order.pharmacy.user else None
        if pharmacy_user and pharmacy_user.is_active:
            Notification.objects.create(
                user=pharmacy_user,
                type='system',
                title='Yetkazib berish holati yangilandi',
                message=f'{delivery.order.order_number} - "{status_labels.get(new_status)}"',
                link=f'/pharmacy/orders/{delivery.order.id}',
            )

        return Response(DeliverySerializer(delivery, context={'request': request}).data)

    @action(detail=True, methods=['get'])
    def location_history(self, request, pk=None):
        delivery = self.get_object()
        logs = delivery.location_logs.all()[:500]
        serializer = DeliveryLocationLogSerializer(logs, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def my_deliveries(self, request):
        qs = self.get_queryset().filter(courier=request.user)
        page = self.paginate_queryset(qs)
        serializer = DeliverySerializer(page or qs, many=True, context={'request': request})
        return self.get_paginated_response(serializer.data) if page else Response(serializer.data)


    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Yetkazib berish'

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        headers = ['Buyurtma', 'Dorixona', 'Kuryer', 'Holati', 'Manzil', 'Yaratilgan vaqt']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        status_labels = dict(Delivery.STATUS_CHOICES)
        for row, d in enumerate(qs, 2):
            data = [
                d.order.order_number if d.order else '-',
                d.order.pharmacy.name if d.order and d.order.pharmacy else '-',
                (d.courier.get_full_name() or d.courier.login) if d.courier else '-',
                status_labels.get(d.status, d.status),
                d.delivery_address or '',
                d.created_at.strftime('%d.%m.%Y %H:%M') if d.created_at else '',
            ]
            for col, val in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = thin_border

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="yetkazib_berish.xlsx"'
        wb.save(response)
        return response
