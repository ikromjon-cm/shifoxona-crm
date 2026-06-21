import openpyxl
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from apps.medicines.models import Medicine, MedicineCategory, Supplier


class Command(BaseCommand):
    help = 'Import medicines/batches from Excel file'

    def add_arguments(self, parser):
        parser.add_argument('file', type=str, help='Excel file path')
        parser.add_argument('--type', type=str, choices=['medicines', 'batches'], default='medicines',
                            help='Import type: medicines or batches')
        parser.add_argument('--dry-run', action='store_true', help='Preview without saving')

    def handle(self, *args, **options):
        file_path = options['file']
        import_type = options['type']
        dry_run = options['dry_run']

        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            ws = wb.active
        except Exception as e:
            raise CommandError(f'File o\'qilmadi: {e}')

        rows = list(ws.iter_rows(min_row=2, values_only=True))
        headers = [cell.value for cell in ws[1]]

        self.stdout.write(f'Fayl: {file_path}')
        self.stdout.write(f'Jami qator: {len(rows)}')

        if import_type == 'medicines':
            self._import_medicines(rows, headers, dry_run)
        elif import_type == 'batches':
            self._import_batches(rows, headers, dry_run)

    def _import_medicines(self, rows, headers, dry_run):
        required = ['name', 'barcode', 'purchase_price', 'selling_price']
        h = {h.lower().strip(): i for i, h in enumerate(headers) if h}

        for req in required:
            if req not in h:
                raise CommandError(f'"{req}" ustuni topilmadi. Mavjud: {", ".join(h.keys())}')

        created = 0
        updated = 0
        errors = []

        for row_idx, row in enumerate(rows, 2):
            try:
                name = str(row[h['name']] or '').strip()
                barcode = str(row[h['barcode']] or '').strip()
                if not name or not barcode:
                    errors.append(f'Qator {row_idx}: nom yoki barcode bo\'sh')
                    continue

                category_name = str(row[h['category']]).strip() if h.get('category') is not None and row[h['category']] else None
                category = None
                if category_name:
                    category, _ = MedicineCategory.objects.get_or_create(name=category_name)

                supplier_name = str(row[h['supplier']]).strip() if h.get('supplier') is not None and row[h['supplier']] else None
                supplier = None
                if supplier_name:
                    supplier, _ = Supplier.objects.get_or_create(
                        name=supplier_name,
                        defaults={'phone': str(row[h.get('supplier_phone', -1)] or '') if h.get('supplier_phone') is not None else ''}
                    )

                purchase_price = float(row[h['purchase_price']] or 0)
                selling_price = float(row[h['selling_price']] or 0)

                data = {
                    'name': name,
                    'category': category,
                    'supplier': supplier,
                    'barcode': barcode,
                    'purchase_price': purchase_price,
                    'selling_price': selling_price,
                    'quantity': int(row[h['quantity']]) if h.get('quantity') is not None and row[h['quantity']] else 0,
                    'min_quantity': int(row[h['min_quantity']]) if h.get('min_quantity') is not None and row[h['min_quantity']] else 10,
                    'description': str(row[h['description']] or '') if h.get('description') is not None else '',
                    'is_active': True,
                }

                if dry_run:
                    self.stdout.write(f'  [DRY] {name} ({barcode}) - import qilinadi')
                    continue

                with transaction.atomic():
                    medicine, was_created = Medicine.objects.update_or_create(
                        barcode=barcode,
                        defaults=data,
                    )
                    if was_created:
                        created += 1
                    else:
                        updated += 1

            except Exception as e:
                errors.append(f'Qator {row_idx}: {e}')

        if not dry_run:
            self.stdout.write(self.style.SUCCESS(
                f'Yaratilgan: {created}, Yangilangan: {updated}, Xatolik: {len(errors)}'
            ))
        if errors:
            for e in errors:
                self.stdout.write(self.style.WARNING(e))

    def _import_batches(self, rows, headers, dry_run):
        h = {h.lower().strip(): i for i, h in enumerate(headers) if h}

        required = ['barcode', 'quantity', 'expiry_date']
        for req in required:
            if req not in h:
                raise CommandError(f'"{req}" ustuni topilmadi')

        created = 0
        errors = []

        for row_idx, row in enumerate(rows, 2):
            try:
                barcode = str(row[h['barcode']] or '').strip()
                if not barcode:
                    errors.append(f'Qator {row_idx}: barcode bo\'sh')
                    continue

                try:
                    medicine = Medicine.objects.get(barcode=barcode, is_active=True)
                except Medicine.DoesNotExist:
                    errors.append(f'Qator {row_idx}: mahsulot topilmadi (barcode: {barcode})')
                    continue

                quantity = int(row[h['quantity']] or 0)
                if quantity <= 0:
                    errors.append(f'Qator {row_idx}: miqdor noto\'g\'ri')
                    continue

                from datetime import datetime
                expiry_raw = row[h['expiry_date']]
                if isinstance(expiry_raw, datetime):
                    expiry_date = expiry_raw.date()
                elif isinstance(expiry_raw, str):
                    expiry_date = datetime.strptime(expiry_raw.strip(), '%Y-%m-%d').date()
                else:
                    errors.append(f'Qator {row_idx}: sana formati noto\'g\'ri')
                    continue

                series = str(row[h['series_number']] or '') if h.get('series_number') is not None else ''
                batch_no = str(row[h['batch_number']] or '') if h.get('batch_number') is not None else ''
                purchase_price = float(row[h['purchase_price']] or 0) if h.get('purchase_price') is not None else 0
                selling_price = float(row[h['selling_price']] or 0) if h.get('selling_price') is not None else 0

                if dry_run:
                    self.stdout.write(f'  [DRY] {medicine.name} - {expiry_date} - {quantity} dona')
                    continue

                from apps.medicines.models import MedicineBatch
                MedicineBatch.objects.create(
                    medicine=medicine,
                    series_number=series or f'IMP-{row_idx}',
                    batch_number=batch_no or None,
                    barcode=barcode,
                    quantity=quantity,
                    purchase_price=purchase_price,
                    selling_price=selling_price,
                    expiry_date=expiry_date,
                )
                created += 1

            except Exception as e:
                errors.append(f'Qator {row_idx}: {e}')

        if not dry_run:
            self.stdout.write(self.style.SUCCESS(f'Yaratilgan batchlar: {created}, Xatolik: {len(errors)}'))
        if errors:
            for e in errors:
                self.stdout.write(self.style.WARNING(e))
