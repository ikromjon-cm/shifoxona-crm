import openpyxl
from django.db.models import F
from django.http import HttpResponse
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import filters, viewsets
from rest_framework.decorators import action
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from apps.accounts.permissions import CanManageMedicines
from apps.warehouse.utils.label_printer import generate_batch_label_pdf, generate_sheet_label_pdf
from apps.warehouse.utils.qr_generator import generate_batch_qr_data, generate_qr_base64

from .models import Medicine, MedicineBatch, MedicineCategory, Supplier
from .serializers import (
    MedicineBatchSerializer,
    MedicineCategorySerializer,
    MedicineCreateSerializer,
    MedicineDetailSerializer,
    MedicineListSerializer,
    SupplierSerializer,
)


class MedicineCategoryViewSet(viewsets.ModelViewSet):
    queryset = MedicineCategory.objects.all()
    serializer_class = MedicineCategorySerializer
    permission_classes = [CanManageMedicines]
    pagination_class = None
    search_fields = ['name', 'description']
    ordering_fields = ['name', 'created_at']


class SupplierViewSet(viewsets.ModelViewSet):
    queryset = Supplier.objects.all()
    serializer_class = SupplierSerializer
    permission_classes = [CanManageMedicines]
    pagination_class = None
    search_fields = ['name', 'contact_person', 'phone', 'email']
    ordering_fields = ['name', 'created_at']


class MedicineViewSet(viewsets.ModelViewSet):
    queryset = Medicine.objects.all()
    permission_classes = [CanManageMedicines]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['category', 'supplier', 'is_active']
    search_fields = ['name', 'barcode', 'series_number', 'description']
    ordering_fields = ['name', 'created_at', 'quantity', 'purchase_price', 'selling_price']

    def get_serializer_class(self):
        if self.action == 'list':
            return MedicineListSerializer
        elif self.action in ['create', 'update', 'partial_update']:
            return MedicineCreateSerializer
        return MedicineDetailSerializer

    def get_queryset(self):
        qs = Medicine.objects.all()
        if self.action == 'list':
            qs = qs.select_related('category', 'supplier')
            low_stock = self.request.query_params.get('low_stock')
            if low_stock and low_stock.lower() == 'true':
                qs = qs.filter(quantity__lte=F('min_quantity'))
        elif self.action == 'retrieve':
            qs = qs.prefetch_related('batches').select_related('category', 'supplier')
        return qs

    def perform_create(self, serializer):
        medicine = serializer.save()
        from django.contrib.auth import get_user_model

        from apps.notifications.models import Notification
        User = get_user_model()
        for user in User.objects.filter(is_active=True, is_blocked=False):
            Notification.objects.create(
                user=user,
                type='medicine',
                title='Yangi mahsulot qo\'shildi',
                message=f'{medicine.name} mahsuloti omborga qo\'shildi',
            )

    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser, FormParser])
    def import_excel(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'Excel fayl talab qilinadi'}, status=400)

        import os
        import tempfile

        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                for chunk in file.chunks():
                    tmp.write(chunk)
                tmp_path = tmp.name

            wb = openpyxl.load_workbook(tmp_path, read_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            headers = [cell.value for cell in ws[1]]
            wb.close()

            h = {str(h or '').lower().strip(): i for i, h in enumerate(headers) if h}
            required = ['name', 'barcode', 'purchase_price', 'selling_price']
            missing = [r for r in required if r not in h]
            if missing:
                os.unlink(tmp_path)
                return Response({'error': f'Ustunlar topilmadi: {", ".join(missing)}'}, status=400)

            from django.db import transaction

            from apps.medicines.models import Medicine, MedicineCategory

            created = 0
            updated = 0
            errors = []

            for row_idx, row in enumerate(rows, 2):
                try:
                    name = str(row[h['name']] or '').strip()
                    barcode = str(row[h['barcode']] or '').strip()
                    if not name or not barcode:
                        errors.append(f'#{row_idx}: nom yoki barcode bo\'sh')
                        continue

                    category = None
                    if h.get('category') is not None and row[h['category']]:
                        cat_name = str(row[h['category']]).strip()
                        category, _ = MedicineCategory.objects.get_or_create(name=cat_name)

                    with transaction.atomic():
                        _, was_created = Medicine.objects.update_or_create(
                            barcode=barcode,
                            defaults={
                                'name': name,
                                'category': category,
                                'purchase_price': float(row[h['purchase_price']] or 0),
                                'selling_price': float(row[h['selling_price']] or 0),
                                'quantity': int(row[h['quantity']]) if h.get('quantity') is not None and row[h['quantity']] else 0,
                                'min_quantity': int(row[h['min_quantity']]) if h.get('min_quantity') is not None and row[h['quantity']] else 10,
                                'is_active': True,
                            },
                        )
                        if was_created:
                            created += 1
                        else:
                            updated += 1
                except Exception as e:
                    errors.append(f'#{row_idx}: {e}')

            os.unlink(tmp_path)
            return Response({
                'message': f'Import yakunlandi: {created} yaratildi, {updated} yangilandi',
                'created': created,
                'updated': updated,
                'errors': errors[:20],
                'total_errors': len(errors),
            })

        except Exception as e:
            return Response({'error': f'Import xatosi: {e}'}, status=400)

    @action(detail=False, methods=['get'])
    def by_barcode(self, request):
        barcode = request.query_params.get('barcode')
        if not barcode:
            return Response({'error': 'Barcode parametri talab qilinadi'}, status=400)
        try:
            medicine = Medicine.objects.get(barcode=barcode, is_active=True)
            serializer = self.get_serializer(medicine)
            return Response(serializer.data)
        except Medicine.DoesNotExist:
            return Response({'error': 'Mahsulot topilmadi'}, status=404)


class MedicineBatchViewSet(viewsets.ModelViewSet):
    queryset = MedicineBatch.objects.all()
    serializer_class = MedicineBatchSerializer
    permission_classes = [CanManageMedicines]
    filterset_fields = ['medicine']
    ordering_fields = ['-created_at']

    @action(detail=True, methods=['get'])
    def qr(self, request, pk=None):
        batch = self.get_object()
        qr_data = generate_batch_qr_data(batch)
        qr_base64 = generate_qr_base64(qr_data)
        return Response({
            'qr_data': qr_data,
            'qr_base64': f'data:image/png;base64,{qr_base64}',
            'batch_id': batch.id,
            'batch_number': batch.batch_number or batch.series_number,
        })

    @action(detail=True, methods=['get'])
    def label(self, request, pk=None):
        batch = self.get_object()
        pdf_buffer = generate_batch_label_pdf(batch)
        return HttpResponse(pdf_buffer.read(), content_type='application/pdf',
                            headers={'Content-Disposition': f'inline; filename="batch_{batch.id}.pdf"'})

    @action(detail=False, methods=['post'])
    def print_labels(self, request):
        ids = request.data.get('ids', [])
        if not ids:
            return Response({'error': 'Batch ID lar talab qilinadi'}, status=400)
        batches = MedicineBatch.objects.filter(id__in=ids)
        pdf_buffer = generate_sheet_label_pdf(batches, 'batch')
        return HttpResponse(pdf_buffer.read(), content_type='application/pdf',
                            headers={'Content-Disposition': 'attachment; filename="batch_labels.pdf"'})
