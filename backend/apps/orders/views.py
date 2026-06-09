from rest_framework import viewsets, status, permissions
from rest_framework.response import Response
from rest_framework.decorators import action
from django.db import transaction
from django.utils import timezone
from django.contrib.auth import get_user_model
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from .models import Order, OrderItem
from .serializers import (
    OrderListSerializer, OrderDetailSerializer, OrderCreateSerializer,
    OrderStatusSerializer, OrderReceiveSerializer
)
from apps.accounts.permissions import IsSuperAdmin
from apps.notifications.models import Notification

User = get_user_model()


class OrderViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]

    def get_serializer_class(self):
        if self.action == 'create':
            return OrderCreateSerializer
        elif self.action in ['list', 'retrieve'] and self.request.user.role == 'pharmacy':
            return OrderListSerializer if self.action == 'list' else OrderDetailSerializer
        elif self.action in ['list', 'retrieve']:
            return OrderListSerializer if self.action == 'list' else OrderDetailSerializer
        return OrderCreateSerializer

    def get_queryset(self):
        user = self.request.user
        qs = Order.objects.select_related('pharmacy', 'created_by')
        if user.role == 'pharmacy':
            if hasattr(user, 'pharmacy_profile'):
                qs = qs.filter(pharmacy=user.pharmacy_profile)
            else:
                qs = qs.none()
        if self.action == 'retrieve':
            qs = qs.prefetch_related('items__medicine', 'delivery')
        return qs

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    @action(detail=True, methods=['post'])
    def update_status(self, request, pk=None):
        serializer = OrderStatusSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        order = self.get_object()
        new_status = serializer.validated_data['status']

        valid_transitions = {
            'pending': ['confirmed', 'cancelled'],
            'confirmed': ['preparing', 'cancelled'],
            'preparing': ['shipped', 'cancelled'],
            'shipped': ['delivered', 'cancelled'],
            'delivered': ['received', 'cancelled'],
        }

        allowed = valid_transitions.get(order.status, [])
        if new_status not in allowed:
            return Response(
                {'error': f'"{order.get_status_display()}" dan "{dict(Order.STATUS_CHOICES).get(new_status)}" ga o\'tish mumkin emas'},
                status=status.HTTP_400_BAD_REQUEST
            )

        with transaction.atomic():
            order.status = new_status
            if new_status == 'received':
                order.received_at = timezone.now()
                order.received_by = request.data.get('received_by', request.user.get_full_name())
                order.receive_note = request.data.get('receive_note', '')
            serializer.validated_data.get('note') and setattr(order, 'note', serializer.validated_data['note'])
            order.save()

        status_labels = dict(Order.STATUS_CHOICES)
        pharmacy_user = order.pharmacy.user if order.pharmacy and order.pharmacy.user else None
        if pharmacy_user and pharmacy_user.is_active:
            Notification.objects.create(
                user=pharmacy_user,
                type='system',
                title='Buyurtma holati yangilandi',
                message=f'{order.order_number} - "{status_labels.get(new_status)}"',
                link=f'/pharmacy/orders/{order.id}',
            )

        return Response(OrderDetailSerializer(order, context={'request': request}).data)

    @action(detail=True, methods=['post'])
    def receive(self, request, pk=None):
        serializer = OrderReceiveSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        order = self.get_object()

        if order.status != 'delivered':
            return Response({'error': 'Faqat "Yetkazilgan" buyurtmalarni qabul qilish mumkin'}, status=status.HTTP_400_BAD_REQUEST)

        with transaction.atomic():
            order.status = 'received'
            order.received_at = timezone.now()
            order.received_by = serializer.validated_data['received_by']
            order.receive_note = serializer.validated_data.get('receive_note', '')

            for item in order.items.all():
                pharmacy_product, _ = order.pharmacy.products.get_or_create(
                    medicine=item.medicine,
                    defaults={'quantity': 0}
                )
                pharmacy_product.quantity += item.quantity
                pharmacy_product.save()

            order.save()

        for user in User.objects.filter(is_active=True, is_blocked=False, role__in=['superadmin', 'operator']):
            Notification.objects.create(
                user=user,
                type='system',
                title='Buyurtma qabul qilindi',
                message=f'{order.order_number} - {order.pharmacy.name} tomonidan qabul qilindi',
                link=f'/warehouse/delivery',
            )

        return Response(OrderDetailSerializer(order, context={'request': request}).data)

    @action(detail=False, methods=['get'])
    def my_orders(self, request):
        if request.user.role != 'pharmacy':
            return Response({'error': 'Faqat dorixonalar uchun'}, status=status.HTTP_403_FORBIDDEN)
        qs = self.get_queryset()
        page = self.paginate_queryset(qs)
        serializer = OrderListSerializer(page or qs, many=True, context={'request': request})
        return self.get_paginated_response(serializer.data) if page else Response(serializer.data)

    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        qs = self.get_queryset().select_related('pharmacy', 'created_by').prefetch_related('items__medicine')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Buyurtmalar'

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        headers = ['Buyurtma raqami', 'Dorixona', 'Holati', 'Mahsulotlar soni', 'Jami summa', 'Izoh', 'Yaratilgan vaqt']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        status_labels = dict(Order.STATUS_CHOICES)
        for row, order in enumerate(qs, 2):
            data = [
                order.order_number,
                order.pharmacy.name if order.pharmacy else '-',
                status_labels.get(order.status, order.status),
                order.total_items,
                float(order.total_amount),
                order.note or '',
                order.created_at.strftime('%d.%m.%Y %H:%M') if order.created_at else '',
            ]
            for col, val in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = thin_border

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="buyurtmalar.xlsx"'
        wb.save(response)
        return response
