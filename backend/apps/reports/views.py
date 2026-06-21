import csv
import io
from datetime import timedelta

from django.db import models
from drf_spectacular.utils import extend_schema
from django.db.models.functions import TruncMonth
from django.http import HttpResponse
from django.utils import timezone
from django.utils.decorators import method_decorator
from django.views.decorators.cache import cache_page
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from rest_framework import generics, status, viewsets
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from apps.accounts.permissions import IsAdmin, IsFinance
from apps.inventory.models import Inventory
from apps.medicines.models import Medicine, MedicineBatch
from apps.pharmacies.models import Pharmacy
from apps.warehouse.models import ExpenseTransaction, IncomeTransaction

from .models import Report
from .serializers import ReportGenerateSerializer, ReportSerializer


class ReportViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Report.objects.all()
    serializer_class = ReportSerializer
    permission_classes = [IsFinance]
    ordering_fields = ['-created_at']

    def get_queryset(self):
        return Report.objects.select_related('created_by')

    @action(detail=False, methods=['post'])
    def generate(self, request):
        serializer = ReportGenerateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        report_type = serializer.validated_data['report_type']
        file_format = serializer.validated_data.get('file_format', 'xlsx')
        params = serializer.validated_data

        if file_format == 'xlsx':
            file_data, filename = self._generate_excel(report_type, params)
            content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        elif file_format == 'csv':
            file_data, filename = self._generate_csv(report_type, params)
            content_type = 'text/csv'
        else:
            return Response({'error': 'Not supported format'}, status=status.HTTP_400_BAD_REQUEST)

        Report.objects.create(
            title=f'{report_type}_{timezone.now().date()}',
            report_type=report_type,
            file_format=file_format,
            filters=params,
            created_by=request.user,
            is_ready=True
        )

        response = HttpResponse(file_data, content_type=content_type)
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    def _generate_excel(self, report_type, params):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Hisobot'

        header_font = Font(bold=True, color='FFFFFF', size=12)
        header_fill = PatternFill(start_color='1A73E8', end_color='1A73E8', fill_type='solid')

        if report_type == 'income':
            data = self._get_income_data(params)
            headers = ['ID', 'Mahsulot', 'Yetkazib beruvchi', 'Miqdor', 'Narx', 'Jami summa', 'Sana', 'Kim qabul qilgan']
            ws.append(headers)
            for row in data:
                ws.append(row)

        elif report_type == 'expense':
            data = self._get_expense_data(params)
            headers = ['ID', 'Mahsulot', 'Dorixona', 'Miqdor', 'Narx', 'Jami summa', 'Sana', 'Kim bergan']
            ws.append(headers)
            for row in data:
                ws.append(row)

        elif report_type == 'inventory':
            data = self._get_inventory_data(params)
            headers = ['ID', 'Mahsulot', 'Barcode', 'Miqdor', 'Minimal miqdor']
            ws.append(headers)
            for row in data:
                ws.append(row)

        elif report_type == 'expiry':
            data = self._get_expiry_data(params)
            headers = ['ID', 'Mahsulot', 'Seriya', 'Miqdor', 'Yaroqlilik muddati']
            ws.append(headers)
            for row in data:
                ws.append(row)

        else:
            ws.append(['Ma\'lumot topilmadi'])

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue(), f'{report_type}_hisobot_{timezone.now().date()}.xlsx'

    def _generate_csv(self, report_type, params):
        output = io.StringIO()
        writer = csv.writer(output)

        if report_type == 'income':
            writer.writerow(['ID', 'Mahsulot', 'Yetkazib beruvchi', 'Miqdor', 'Narx', 'Jami summa', 'Sana', 'Kim qabul qilgan'])
            for row in self._get_income_data(params):
                writer.writerow(row)
        elif report_type == 'expense':
            writer.writerow(['ID', 'Mahsulot', 'Dorixona', 'Miqdor', 'Narx', 'Jami summa', 'Sana', 'Kim bergan'])
            for row in self._get_expense_data(params):
                writer.writerow(row)
        else:
            writer.writerow(['Ma\'lumot topilmadi'])

        output.seek(0)
        return output.getvalue().encode('utf-8'), f'{report_type}_hisobot_{timezone.now().date()}.csv'

    def _get_income_data(self, params):
        qs = IncomeTransaction.objects.select_related('medicine', 'supplier', 'created_by')
        if params.get('start_date'):
            qs = qs.filter(created_at__date__gte=params['start_date'])
        if params.get('end_date'):
            qs = qs.filter(created_at__date__lte=params['end_date'])
        data = []
        for t in qs:
            data.append([
                t.id, t.medicine.name,
                t.supplier.name if t.supplier else '-',
                t.quantity, float(t.price), float(t.total_amount),
                t.created_at.strftime('%Y-%m-%d %H:%M'),
                f"{t.created_by.first_name} {t.created_by.last_name}" if t.created_by else '-'
            ])
        return data

    def _get_expense_data(self, params):
        qs = ExpenseTransaction.objects.select_related('medicine', 'pharmacy', 'created_by')
        if params.get('start_date'):
            qs = qs.filter(created_at__date__gte=params['start_date'])
        if params.get('end_date'):
            qs = qs.filter(created_at__date__lte=params['end_date'])
        data = []
        for t in qs:
            data.append([
                t.id, t.medicine.name,
                t.pharmacy.name if t.pharmacy else '-',
                t.quantity, float(t.price), float(t.total_amount),
                t.created_at.strftime('%Y-%m-%d %H:%M'),
                f"{t.created_by.first_name} {t.created_by.last_name}" if t.created_by else '-'
            ])
        return data

    def _get_inventory_data(self, params):
        qs = Inventory.objects.select_related('medicine').all()
        data = []
        for inv in qs:
            data.append([inv.id, inv.medicine.name, inv.medicine.barcode, inv.quantity, inv.min_quantity])
        return data

    def _get_expiry_data(self, params):
        qs = MedicineBatch.objects.filter(quantity__gt=0).select_related('medicine')
        data = []
        for batch in qs:
            data.append([batch.id, batch.medicine.name, batch.series_number, batch.quantity, batch.expiry_date])
        return data


@extend_schema(exclude=True)
class DashboardView(generics.GenericAPIView):
    permission_classes = [IsAuthenticated]
    cache_key_prefix = 'dashboard_main'

    @method_decorator(cache_page(120, key_prefix='dashboard_main'))
    def get(self, request):
        total_medicines = Medicine.objects.count()
        total_quantity = Medicine.objects.aggregate(models.Sum('quantity'))['quantity__sum'] or 0
        today = timezone.now().date()
        today_income = IncomeTransaction.objects.filter(
            created_at__date=today
        ).aggregate(total=models.Sum('total_amount'))['total'] or 0
        today_expense = ExpenseTransaction.objects.filter(
            created_at__date=today
        ).aggregate(total=models.Sum('total_amount'))['total'] or 0
        low_stock = Medicine.objects.filter(quantity__lte=models.F('min_quantity')).count()
        total_pharmacies = Pharmacy.objects.filter(is_active=True).count()

        expiring_soon = MedicineBatch.objects.filter(
            expiry_date__lte=timezone.now().date() + timedelta(days=30),
            quantity__gt=0
        ).count()

        top_medicines = ExpenseTransaction.objects.values(
            'medicine__name'
        ).annotate(
            total_qty=models.Sum('quantity'),
            total_amount=models.Sum('total_amount')
        ).order_by('-total_qty')[:10]

        monthly_income = IncomeTransaction.objects.filter(
            created_at__year=timezone.now().year
        ).annotate(
            month=TruncMonth('created_at')
        ).values('month').annotate(
            total=models.Sum('total_amount'),
            count=models.Count('id')
        ).order_by('month')

        monthly_expense = ExpenseTransaction.objects.filter(
            created_at__year=timezone.now().year
        ).annotate(
            month=TruncMonth('created_at')
        ).values('month').annotate(
            total=models.Sum('total_amount'),
            count=models.Count('id')
        ).order_by('month')

        from apps.orders.models import Order
        today_orders = Order.objects.filter(created_at__date=today).count()
        pending_orders = Order.objects.filter(status='pending').count()
        delivered_orders = Order.objects.filter(status='delivered').count()
        received_orders = Order.objects.filter(status='received').count()

        top_pharmacies = Order.objects.values(
            'pharmacy__name', 'pharmacy__id'
        ).annotate(
            total_orders=models.Count('id'),
            total_amount=models.Sum('total_amount')
        ).order_by('-total_orders')[:10]

        pharmacy_locations = Pharmacy.objects.filter(
            is_active=True, latitude__isnull=False, longitude__isnull=False
        ).values('id', 'name', 'latitude', 'longitude', 'address', 'phone')

        return Response({
            'total_medicines': total_medicines,
            'total_quantity': total_quantity,
            'today_income': today_income,
            'today_expense': today_expense,
            'low_stock': low_stock,
            'total_pharmacies': total_pharmacies,
            'expiring_soon': expiring_soon,
            'top_medicines': list(top_medicines),
            'monthly_income': list(monthly_income),
            'monthly_expense': list(monthly_expense),
            'total_pharmacies_active': Pharmacy.objects.filter(is_active=True, is_approved=True).count(),
            'today_orders': today_orders,
            'pending_orders': pending_orders,
            'delivered_orders': delivered_orders,
            'received_orders': received_orders,
            'top_pharmacies': list(top_pharmacies),
            'pharmacy_locations': list(pharmacy_locations),
        })


@extend_schema(exclude=True)
class AdminDashboardView(generics.GenericAPIView):
    permission_classes = [IsAuthenticated]

    @method_decorator(cache_page(120, key_prefix='dashboard_admin'))
    def get(self, request):
        from django.contrib.auth import get_user_model

        from apps.attendance.models import AttendanceSession, LeaveRequest
        from apps.chat.models import ChatMessage, ChatRoom
        from apps.delivery.models import Delivery
        from apps.orders.models import Order
        from apps.tasks.models import Task
        from apps.warehouse.models import PickOrder, PickWave, Warehouse

        User = get_user_model()
        today = timezone.now().date()

        # Users
        users = User.objects.all()
        user_stats = {
            'total': users.count(),
            'active': users.filter(is_active=True, is_blocked=False).count(),
            'blocked': users.filter(is_blocked=True).count(),
            'by_role': {
                role: users.filter(role=role).count()
                for role, _ in User.ROLE_CHOICES
            },
        }

        # Warehouse
        warehouse_stats = {
            'total_warehouses': Warehouse.objects.count(),
            'active_pick_waves': PickWave.objects.filter(status='in_progress').count(),
            'pending_pick_orders': PickOrder.objects.filter(status='pending').count(),
            'completed_today_picks': PickOrder.objects.filter(
                completed_at__date=today, status='picked'
            ).count(),
        }

        # Orders
        order_stats = {
            'total': Order.objects.count(),
            'today': Order.objects.filter(created_at__date=today).count(),
            'pending': Order.objects.filter(status='pending').count(),
            'in_progress': Order.objects.filter(status__in=['confirmed', 'preparing']).count(),
            'shipped': Order.objects.filter(status='shipped').count(),
            'delivered': Order.objects.filter(status='delivered').count(),
            'received': Order.objects.filter(status='received').count(),
            'cancelled': Order.objects.filter(status='cancelled').count(),
        }

        # Delivery
        delivery_stats = {
            'pending': Delivery.objects.filter(status='pending').count(),
            'in_transit': Delivery.objects.filter(status='in_transit').count(),
            'delivered_today': Delivery.objects.filter(
                delivered_at__date=today, status='delivered'
            ).count(),
            'active_couriers': User.objects.filter(
                role='operator', is_active=True, is_blocked=False
            ).count(),
        }

        # Attendance
        attendance_stats = {
            'checked_in_today': AttendanceSession.objects.filter(
                date=today, check_in__isnull=False
            ).count(),
            'on_leave_today': LeaveRequest.objects.filter(
                status='approved',
                start_date__lte=today,
                end_date__gte=today,
            ).count(),
            'pending_leaves': LeaveRequest.objects.filter(status='pending').count(),
        }

        # Tasks
        task_stats = {
            'total': Task.objects.count(),
            'pending': Task.objects.filter(status='pending').count(),
            'in_progress': Task.objects.filter(status='in_progress').count(),
            'completed_today': Task.objects.filter(
                completed_at__date=today, status='completed'
            ).count(),
            'overdue': Task.objects.filter(
                status__in=['pending', 'in_progress'],
                due_date__lt=today,
            ).count(),
        }

        # Chat
        chat_stats = {
            'total_rooms': ChatRoom.objects.count(),
            'active_today': ChatRoom.objects.filter(
                updated_at__date=today
            ).count(),
            'messages_today': ChatMessage.objects.filter(
                created_at__date=today
            ).count(),
        }

        # Revenue
        from apps.warehouse.models import ExpenseTransaction, IncomeTransaction
        this_month_start = today.replace(day=1)
        revenue_stats = {
            'today_income': (IncomeTransaction.objects.filter(
                created_at__date=today
            ).aggregate(t=models.Sum('total_amount'))['t'] or 0),
            'today_expense': (ExpenseTransaction.objects.filter(
                created_at__date=today
            ).aggregate(t=models.Sum('total_amount'))['t'] or 0),
            'month_income': (IncomeTransaction.objects.filter(
                created_at__date__gte=this_month_start
            ).aggregate(t=models.Sum('total_amount'))['t'] or 0),
            'month_expense': (ExpenseTransaction.objects.filter(
                created_at__date__gte=this_month_start
            ).aggregate(t=models.Sum('total_amount'))['t'] or 0),
        }

        # Monthly trend (last 6 months)
        six_months_ago = today - timedelta(days=180)
        monthly_income = IncomeTransaction.objects.filter(
            created_at__date__gte=six_months_ago
        ).annotate(
            month=TruncMonth('created_at')
        ).values('month').annotate(
            total=models.Sum('total_amount'),
            count=models.Count('id')
        ).order_by('month')

        monthly_expense = ExpenseTransaction.objects.filter(
            created_at__date__gte=six_months_ago
        ).annotate(
            month=TruncMonth('created_at')
        ).values('month').annotate(
            total=models.Sum('total_amount'),
            count=models.Count('id')
        ).order_by('month')

        monthly_orders = Order.objects.filter(
            created_at__date__gte=six_months_ago
        ).annotate(
            month=TruncMonth('created_at')
        ).values('month').annotate(
            count=models.Count('id'),
            total_amount=models.Sum('total_amount')
        ).order_by('month')

        # Inventory alerts
        low_stock = Medicine.objects.filter(
            quantity__lte=models.F('min_quantity'), is_active=True
        ).values('id', 'name', 'quantity', 'min_quantity')[:20]

        expiring_batches = MedicineBatch.objects.filter(
            expiry_date__lte=today + timedelta(days=30),
            expiry_date__gte=today,
            quantity__gt=0,
        ).select_related('medicine').order_by('expiry_date')[:20]

        return Response({
            'users': user_stats,
            'warehouse': warehouse_stats,
            'orders': order_stats,
            'delivery': delivery_stats,
            'attendance': attendance_stats,
            'tasks': task_stats,
            'chat': chat_stats,
            'revenue': revenue_stats,
            'trends': {
                'monthly_income': [
                    {'month': m['month'], 'total': float(m['total']), 'count': m['count']}
                    for m in monthly_income
                ],
                'monthly_expense': [
                    {'month': m['month'], 'total': float(m['total']), 'count': m['count']}
                    for m in monthly_expense
                ],
                'monthly_orders': [
                    {'month': m['month'], 'count': m['count'], 'total_amount': float(m['total_amount'])}
                    for m in monthly_orders
                ],
            },
            'alerts': {
                'low_stock': list(low_stock),
                'expiring_batches': [
                    {
                        'id': b.id,
                        'medicine_name': b.medicine.name,
                        'batch': b.batch_number or b.series_number,
                        'expiry_date': b.expiry_date,
                        'quantity': b.quantity,
                    }
                    for b in expiring_batches
                ],
            },
        })
